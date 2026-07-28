import io, os, datetime as dt, requests, pandas as pd, streamlit as st, altair as alt
from concurrent.futures import ThreadPoolExecutor
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from bs4 import BeautifulSoup

st.set_page_config(page_title="NSE Derivatives", page_icon="📊", layout="wide")
REF = "https://www.nseindia.com/all-reports-derivatives"
ARCH = "https://nsearchives.nseindia.com/content/nsccl"
PIT_REF = "https://www.nseindia.com/companies-listing/corporate-filings-insider-trading"
PIT_API = "https://www.nseindia.com/api/corporates-pit-gg"
FPI_URL = "https://www.fpi.nsdl.co.in/web/Reports/FPI_Fortnightly_Selection.aspx"
FPI_BASE = "https://www.fpi.nsdl.co.in/web/"
PARTS = ["FII", "Pro", "Client", "DII"]
NAVY, GREEN = "#1F3864", "#2E9E5B"
LOGO = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Ashika_logo-removebg-preview.png")

GROUPS = [
    ("Sr. No.", None), ("Type of instrument", None),
    ("Description of type of instrument (applicable in case of other is selected)", None),
    ("Category of person", None), ("Name of the person", None), ("CIN / DIN", None),
    ("Securities held prior to acquisition / disposal", "No. of security"),
    ("Securities held prior to acquisition / disposal", "% of shareholding"),
    ("Securities acquired / disposed", "No. of security"),
    ("Securities acquired / disposed", "Value of security (in Rs.)"),
    ("Securities acquired / disposed", "Transaction type"),
    ("Securities held post acquisition / disposal", "No. of security"),
    ("Securities held post acquisition / disposal", "% of shareholding"),
    ("Date of allotment advice / acquisition of shares / sale of shares specify", "From date"),
    ("Date of allotment advice / acquisition of shares / sale of shares specify", "To date"),
    ("Mode of acquisition / disposal", None), ("Date of intimation to company", None),
    ("Type of contract", None), ("Contract specification", None),
    ("Buy", "Notional value"), ("Buy", "No. of units (contracts x lot size)"),
    ("Sell", "Notional value"), ("Sell", "No. of units (contracts x lot size)"),
    ("Exchange on which the trade was executed", None), ("Notes", None),
]
HEAD = [("Symbol", None), ("Company", None), ("Broadcast", None),
        ("Regulation", None), ("Submission", None)] + GROUPS
FLAT = [t if s is None else f"{t} - {s}" for t, s in HEAD]

PIT_HIDE = {"Sr. No.", "Category of person", "CIN / DIN",
            "Description of type of instrument (applicable in case of other is selected)"}
PIT_FRONT = ["Name of the person", "Type of instrument"]
PIT_VIEW = (["Broadcast", "Regulation", "Submission"] + PIT_FRONT +
            [c for c in FLAT if c not in PIT_HIDE and
             c not in PIT_FRONT + ["Symbol", "Company", "Broadcast", "Regulation", "Submission"]])


def view_cols(df, keep_symbol=False):
    cols = (["Symbol", "Company"] if keep_symbol else []) + PIT_VIEW
    return df[[c for c in cols if c in df.columns]]


@st.cache_resource
def session():
    s = requests.Session()
    s.headers.update({"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                                    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0 Safari/537.36",
                      "Referer": REF})
    s.get(REF, timeout=15)
    return s


def get(url, **kw):
    try:
        r = session().get(url, timeout=30, **kw)
        return r.content if r.ok and len(r.content) > 200 else None
    except Exception:
        return None


@st.cache_data(ttl=300, show_spinner=False)
def load_mwpl():
    d = dt.date.today()
    for _ in range(10):
        c = get(f"{ARCH}/mwpl_cli_{d:%d%m%Y}.xls")
        if c and len(c) > 500:
            df = pd.read_excel(io.BytesIO(c), skiprows=1)
            cli = df.columns[2:]
            df[cli] = df[cli].apply(pd.to_numeric, errors="coerce")
            df["Count"] = df[cli].count(axis=1)
            df["Sum"] = df[cli].sum(axis=1).round(2)
            df["Average"] = df[cli].mean(axis=1).round(2)
            return d, df
        d -= dt.timedelta(days=1)
    return None, None


@st.cache_data(ttl=300, show_spinner=False)
def load_oi(days=30):
    today = dt.date.today()
    dates = [d for d in (today - dt.timedelta(days=i) for i in range(days + 1)) if d.weekday() < 5]

    def one(d):
        c = get(f"{ARCH}/fao_participant_oi_{d:%d%m%Y}.csv")
        if not c:
            return None
        df = pd.read_csv(io.BytesIO(c), skiprows=1)
        df.columns = [x.strip() for x in df.columns]
        df["Client Type"] = df["Client Type"].astype(str).str.strip()
        df = df.set_index("Client Type")
        row = {"_d": d}
        for p in PARTS:
            if p not in df.index:
                return None
            lo, sh = float(df.loc[p, "Future Index Long"]), float(df.loc[p, "Future Index Short"])
            t = lo + sh
            row[f"{p} Long %"] = round(lo / t * 100, 1) if t else 0.0
            row[f"{p} Short %"] = round(sh / t * 100, 1) if t else 0.0
            row[f"{p} Total"] = int(t)
        return row

    with ThreadPoolExecutor(max_workers=5) as ex:
        rows = [r for r in ex.map(one, dates) if r]
    if not rows:
        return None
    out = pd.DataFrame(rows).sort_values("_d", ascending=False)
    out["_d"] = out["_d"].map(lambda x: x.strftime("%d-%m-%Y"))
    return out.rename(columns={"_d": "Dates"}).reset_index(drop=True)


@st.cache_data(ttl=300, show_spinner=False)
def load_pit(from_d, to_d, index="equities"):
    s = session()
    try:
        s.get(PIT_REF, timeout=15)
        s.get("https://www.nseindia.com/api/marketStatus", timeout=15,
              headers={"Referer": PIT_REF})
        r = s.get(PIT_API, params={"index": index,
                                   "from_date": from_d.strftime("%d-%m-%Y"),
                                   "to_date": to_d.strftime("%d-%m-%Y")},
                  timeout=25, headers={"Referer": PIT_REF})
        r.raise_for_status()
        j = r.json()
        rows = j["data"] if isinstance(j, dict) else j
    except Exception:
        return None, None, [("-", "NSE API unreachable")]

    rows = [x for x in rows if x.get("ixbrl")]
    if not rows:
        return pd.DataFrame(), pd.DataFrame(columns=FLAT), []

    def one(row):
        try:
            c = get(row["ixbrl"], headers={"Referer": PIT_REF})
            if not c:
                return row.get("symbol"), None, None
            tables = pd.read_html(io.StringIO(c.decode("utf-8", errors="replace")))
            inf = tables[0][[0, 2]]
            meta = dict(zip(inf[0], inf[2]))
            meta["Symbol"] = row.get("symbol")
            meta["Company"] = row.get("companyName")
            pit = tables[1]
            pit.columns = range(pit.shape[1])
            pre = [row.get("symbol"), row.get("companyName"), row.get("broadcastDateTime"),
                   row.get("regulation"), row.get("typeOfSubmission")]
            return row.get("symbol"), meta, [pre + list(x) for x in pit.itertuples(index=False)]
        except Exception as e:
            return row.get("symbol"), None, str(e)[:120]

    metas, recs, failed = [], [], []
    with ThreadPoolExecutor(max_workers=5) as ex:
        for sym, meta, out in ex.map(one, rows):
            if meta is None:
                failed.append((sym, out if isinstance(out, str) else "download failed"))
            else:
                metas.append(meta)
                recs.extend(out)

    inf_df = pd.DataFrame(metas)
    if not inf_df.empty:
        inf_df = inf_df[["Symbol", "Company"] +
                        [c for c in inf_df.columns if c not in ("Symbol", "Company")]]
    return inf_df, pd.DataFrame(recs, columns=FLAT), failed


def _fpi_session():
    fs = requests.Session()
    fs.headers.update({"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                                     "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0 Safari/537.36"})
    return fs


def _fpi_num(t):
    t = t.replace(",", "").strip()
    try:
        return int(t)
    except ValueError:
        try:
            return float(t)
        except ValueError:
            return None


def _fpi_parse(fs, val):
    grid = [[(c.get_text(strip=True), int(c.get("colspan") or 1))
             for c in tr.find_all(["td", "th"])]
            for tr in BeautifulSoup(fs.get(FPI_BASE + val.replace("~/", ""), timeout=30).text,
                                    "html.parser").find_all("table")[0].find_all("tr")]
    exp = lambda r: [t for t, sp in r for _ in range(sp)]
    b, u, sb = exp(grid[0]), exp(grid[1]), exp(grid[3])
    cols = {b[ci]: ci for ci in range(2, 98)
            if "Net Investment" in b[ci] and u[ci] == "IN INR Cr." and sb[ci] == "Total"}
    return {p: {grid[ri][1][0]: _fpi_num(grid[ri][ci][0])
                for ri in range(4, len(grid)) if grid[ri][1][0]}
            for p, ci in cols.items()}


@st.cache_data(ttl=86400, show_spinner=False)
def fpi_options():
    try:
        fs = _fpi_session()
        return [(o["value"], o.text.strip())
                for o in BeautifulSoup(fs.get(FPI_URL, timeout=30).text, "html.parser")
                .select("select[name=ddlfortnighly] option") if o.get("value")]
    except Exception:
        return None


@st.cache_data(ttl=3600, show_spinner=False)
def fpi_fetch(vals):
    fs = _fpi_session()
    periods = {}
    for val in vals:
        for p, d in _fpi_parse(fs, val).items():
            periods.setdefault(p, d)
    return periods


MON = dict(zip("Jan Feb Mar Apr May Jun Jul Aug Sep Oct Nov Dec".split(), range(1, 13)))


def fpi_key(p):
    l, y = p.replace("Net Investment ", "").rsplit(", ", 1)
    return (int(y), MON[l[:3]], int(l.split("-")[-1].split()[0]))


def fpi_month(p):
    l, y = p.replace("Net Investment ", "").rsplit(", ", 1)
    return f"{l[:3]} {y}"


def pit_xlsx(inf_df, pit_df, failed):
    thin = Border(*[Side(style="thin", color="999999")] * 4)
    hdr = PatternFill("solid", fgColor="D9D9D9")
    ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wb = Workbook()

    ws = wb.active
    ws.title = "Filings"
    for i, c in enumerate(inf_df.columns, 1):
        cell = ws.cell(1, i, c)
        cell.font = Font(bold=True); cell.fill = hdr; cell.border = thin; cell.alignment = ctr
        ws.column_dimensions[get_column_letter(i)].width = 24
    for ri, rec in enumerate(inf_df.itertuples(index=False), 2):
        for ci, v in enumerate(rec, 1):
            ws.cell(ri, ci, "" if pd.isna(v) else v).border = thin
    ws.freeze_panes = "C2"

    ws2 = wb.create_sheet("PIT Disclosure")
    for i, (top, sub) in enumerate(HEAD, 1):
        ws2.cell(1, i, top)
        if sub is None:
            ws2.merge_cells(start_row=1, start_column=i, end_row=2, end_column=i)
        else:
            ws2.cell(2, i, sub)
    for top in dict.fromkeys(t for t, sub in HEAD if sub is not None):
        idx = [i for i, (t, sub) in enumerate(HEAD, 1) if t == top]
        ws2.merge_cells(start_row=1, start_column=min(idx), end_row=1, end_column=max(idx))
    for hr in (1, 2):
        for i in range(1, len(HEAD) + 1):
            c = ws2.cell(hr, i)
            c.font = Font(bold=True); c.fill = hdr; c.alignment = ctr; c.border = thin
    for ri, rec in enumerate(pit_df.itertuples(index=False), 3):
        for ci, v in enumerate(rec, 1):
            c = ws2.cell(ri, ci, "" if pd.isna(v) else v)
            c.border = thin
            c.alignment = Alignment(vertical="center", wrap_text=True)
    for i in range(1, len(HEAD) + 1):
        ws2.column_dimensions[get_column_letter(i)].width = 20
    ws2.row_dimensions[1].height = 45
    ws2.row_dimensions[2].height = 30
    ws2.freeze_panes = "C3"

    if failed:
        ws3 = wb.create_sheet("Failed")
        ws3.append(["Symbol", "Error"])
        for f in failed:
            ws3.append(list(f))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def xl(df, sheet):
    return df.to_csv(index=False).encode()


st.markdown("""
<style>
  div[data-testid="stButton"] button {padding:2px 12px; font-size:12px; min-height:0;}
  div[data-testid="stButton"] {display:flex; justify-content:flex-end;}
  h1 {padding-top:0 !important; margin-top:0 !important;}
</style>
""", unsafe_allow_html=True)

a1, a2 = st.columns([5, 1], vertical_alignment="center")
with a1:
    st.title("NSE Derivatives Monitor")
with a2:
    if os.path.exists(LOGO):
        st.image(LOGO, width=130)

b1, b2 = st.columns([5, 1], vertical_alignment="center")
with b1:
    st.caption("MWPL client positions · participant-wise open interest · insider trading · FPI flows · live")
with b2:
    if st.button("Refresh"):
        load_mwpl.clear()
        load_oi.clear()
        load_pit.clear()
        fpi_options.clear()
        fpi_fetch.clear()

t0, t1, t2, t3, t4 = st.tabs(["About", "MWPL Client Positions",
                              "Participant OI (1M)", "Insider Trading (PIT)",
                              "FPI Sector Flows"])

with t0:
    st.markdown("""
**MWPL Client Positions**  
NSE's "F&O Clients Position % greater than or equal to 3% of Stock MWPL" file.
Each row is an underlying stock; each Client column is one client holding at least
3% of that stock's market-wide position limit.

- **Count**: how many clients are above the 3% threshold
- **Sum**: their combined position as % of MWPL
- **Average**: mean position size per reporting client

High Count with low Average is a crowded trade. Low Count with high Average is
concentration in one or two hands.

**Participant OI**  
NSE's "Participant wise Open Interest" file for the last ~21 trading days.
Long % and Short % are each participant's share of index futures open interest,
computed as long / (long + short). Dashed lines are the average across the window.

**Insider Trading (PIT)**  
Reg 7(2) and 7(3) disclosures under SEBI (Prohibition of Insider Trading) Regulations, 2015.
Each filing is an iXBRL document parsed into two tables: the company's general information
and the PIT disclosure rows (person, category, holdings before and after, quantity, value,
transaction type, mode). Buy/Sell and contract columns are populated only for derivatives.

**FPI Sector Flows**  
NSDL fortnightly sector-wise FII/FPI net investment (INR Cr). Bar chart shows one
fortnight across all sectors; line chart tracks a single sector across a chosen range.

**Notes**  
Position date runs one trading day behind the publish date. Missing days are holidays.
Data refreshes every 5 minutes, or immediately via Refresh. Nothing is stored locally.
Source: NSE India / NSDL. Internal research use only, not investment advice.
""")

with t1:
    date, df = load_mwpl()
    if df is None:
        st.error("No MWPL file found in the last 10 days.")
    else:
        st.caption(f"Position date: {date:%d %b %Y}")
        sym = df.columns[1]
        top = df[[sym, "Sum", "Count", "Average"]].dropna().nlargest(20, "Sum")
        st.subheader("Top 20 by total client position")
        st.altair_chart(
            alt.Chart(top).mark_bar().encode(
                y=alt.Y(f"{sym}:N", sort="-x", title=None),
                x=alt.X("Sum:Q", title="Total % of MWPL"),
                color=alt.Color("Count:Q", scale=alt.Scale(scheme="tealblues"), title="Clients"),
                tooltip=[sym, "Sum", "Count", "Average"],
            ).properties(height=480), use_container_width=True)

        st.dataframe(df, use_container_width=True, hide_index=True,
                     height=min(1200, 35 * len(df) + 40))
        st.download_button("CSV", xl(df, "MWPL"),
                           f"mwpl_cli_{date:%d%m%Y}.csv", "text/csv", key="dl_mwpl")

with t2:
    oi = load_oi()
    if oi is None:
        st.error("No participant OI files found.")
    else:
        st.caption(f"{len(oi)} trading days · latest {oi['Dates'].iloc[0]}")
        src = oi.iloc[::-1].reset_index(drop=True)
        order = list(src["Dates"])
        xax = alt.X("Dates:O", sort=order, title=None,
                    axis=alt.Axis(labelAngle=-45, labelOverlap=True))
        shade = alt.Scale(domain=["Long", "Short"], range=["#26a65b", "#e05260"])
        cols = st.columns(2)
        for i, p in enumerate(PARTS):
            m = src[["Dates", f"{p} Long %", f"{p} Short %"]].melt(
                "Dates", var_name="Side", value_name="Value")
            m["Side"] = m["Side"].str.extract(r"(Long|Short)")
            avg = m.groupby("Side")["Value"].mean().round(2)
            a = pd.DataFrame({"Side": avg.index, "Avg": avg.values})
            with cols[i % 2]:
                st.subheader(p)
                st.caption(f"Avg long {avg['Long']}% · avg short {avg['Short']}%")
                st.altair_chart(
                    (alt.Chart(m).mark_line(strokeWidth=2).encode(
                        x=xax,
                        y=alt.Y("Value:Q", scale=alt.Scale(zero=False), title="% of futures OI"),
                        color=alt.Color("Side:N", scale=shade)) +
                     alt.Chart(m).mark_point(size=55, filled=True).encode(
                         x=xax, y="Value:Q", color=alt.Color("Side:N", scale=shade),
                         tooltip=["Dates", "Side", "Value"]) +
                     alt.Chart(a).mark_rule(strokeDash=[5, 3], opacity=.6).encode(
                         y="Avg:Q", color=alt.Color("Side:N", scale=shade),
                         tooltip=["Side", "Avg"])
                     ).properties(height=320).interactive(), use_container_width=True)

        st.dataframe(oi, use_container_width=True, hide_index=True,
                     height=min(1200, 35 * len(oi) + 40))
        st.download_button("CSV", xl(oi, "Participant OI"),
                           f"participant_oi_{dt.date.today():%d%m%Y}.csv", "text/csv", key="dl_oi")

with t3:
    today = dt.date.today()
    c1, c2, c3 = st.columns([2, 2, 2], vertical_alignment="bottom")
    rng = c1.selectbox("Range", ["1D", "1W", "1M", "Custom"], key="pit_rng")
    seg = c2.selectbox("Segment", ["equities", "sme", "debt", "invitsreits"], key="pit_seg")
    if rng == "Custom":
        d1, d2 = c3.columns(2)
        from_d = d1.date_input("From", today - dt.timedelta(days=7), key="pit_from")
        to_d = d2.date_input("To", today, key="pit_to")
    else:
        back = {"1D": 0, "1W": 7, "1M": 30}[rng]
        from_d, to_d = today - dt.timedelta(days=back), today
        c3.caption(f"{from_d:%d %b} \u2192 {to_d:%d %b %Y}")

    with st.spinner("Fetching filings from NSE..."):
        inf_df, pit_df, failed = load_pit(from_d, to_d, seg)

    if inf_df is None:
        st.error("NSE API unreachable. Hit Refresh and try again.")
    elif inf_df.empty:
        st.warning("No filings in that range.")
    else:
        m1, m2, m3, m4 = st.columns([1, 1, 1, 2], vertical_alignment="bottom")
        m1.metric("Filings", len(inf_df))
        m2.metric("Disclosure rows", len(pit_df))
        m3.metric("Failed", len(failed))
        show_chart = m4.toggle("Show chart", value=True, key="pit_chart")

        buy = pit_df["Securities acquired / disposed - Transaction type"].astype(str)
        val = pd.to_numeric(
            pit_df["Securities acquired / disposed - Value of security (in Rs.)"],
            errors="coerce").fillna(0)
        agg = (pd.DataFrame({"Symbol": pit_df["Symbol"], "Side": buy, "Value": val})
               .query("Side in ['Buy', 'Sell']")
               .groupby(["Symbol", "Side"], as_index=False)["Value"].sum())
        if show_chart and not agg.empty:
            st.subheader("Top 20 by disclosed value")
            top = (agg.groupby("Symbol")["Value"].sum().nlargest(20).index)
            st.altair_chart(
                alt.Chart(agg[agg["Symbol"].isin(top)]).mark_bar().encode(
                    y=alt.Y("Symbol:N", sort="-x", title=None),
                    x=alt.X("Value:Q", title="Value (Rs.)"),
                    color=alt.Color("Side:N", scale=alt.Scale(
                        domain=["Buy", "Sell"], range=["#26a65b", "#e05260"])),
                    tooltip=["Symbol", "Side", "Value"],
                ).properties(height=480), use_container_width=True)

        d1, d2 = st.columns([1, 6])
        d1.download_button("Excel", pit_xlsx(inf_df, pit_df, failed),
                           f"PIT_{seg}_{from_d:%d%m%Y}_{to_d:%d%m%Y}.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key="dl_pit_x")
        d2.download_button("CSV", xl(pit_df, "PIT"),
                           f"PIT_{seg}_{from_d:%d%m%Y}_{to_d:%d%m%Y}.csv",
                           "text/csv", key="dl_pit_c")

        s1, s2 = st.tabs(["By company", "All rows"])

        with s1:
            work = pit_df.copy()
            work["_val"] = val
            work["_side"] = buy
            summary = (work.groupby(["Symbol", "Company"], as_index=False)
                       .agg(Rows=("Symbol", "size"),
                            Filings=("Broadcast", "nunique"),
                            Buy=("_val", lambda x: x[work.loc[x.index, "_side"] == "Buy"].sum()),
                            Sell=("_val", lambda x: x[work.loc[x.index, "_side"] == "Sell"].sum())))
            summary["Net"] = summary["Buy"] - summary["Sell"]
            summary = summary.sort_values("Rows", ascending=False).reset_index(drop=True)

            st.caption(f"{len(summary)} companies \u00b7 click a row to open it below")
            picked = None
            try:
                sel = st.dataframe(summary, use_container_width=True, hide_index=True,
                                   on_select="rerun", selection_mode="single-row",
                                   key="pit_pick", height=min(600, 35 * len(summary) + 40))
                if sel.selection.rows:
                    picked = summary.iloc[sel.selection.rows[0]]
            except TypeError:
                st.dataframe(summary, use_container_width=True, hide_index=True)
                sym = st.selectbox("Open company", ["-"] + list(summary["Symbol"]),
                                   key="pit_pick_fb")
                if sym != "-":
                    picked = summary[summary["Symbol"] == sym].iloc[0]

            if picked is not None:
                st.divider()
                st.subheader(f"{picked.Symbol} \u2014 {picked.Company}")
                st.caption(f"{picked.Rows} disclosure rows across {picked.Filings} filing"
                           f"{'s' if picked.Filings != 1 else ''} \u00b7 "
                           f"buy \u20b9{picked.Buy:,.0f} \u00b7 sell \u20b9{picked.Sell:,.0f} \u00b7 "
                           f"net \u20b9{picked.Net:,.0f}")
                sub = view_cols(pit_df[pit_df["Symbol"] == picked.Symbol])
                st.dataframe(sub, use_container_width=True, hide_index=True,
                             height=min(1200, 35 * len(sub) + 45))

        flat = view_cols(pit_df, keep_symbol=True)
        s2.dataframe(flat, use_container_width=True, hide_index=True,
                     height=min(1200, 35 * len(flat) + 40))

with t4:
    opts = fpi_options()
    if not opts:
        st.error("NSDL unreachable. Hit Refresh and try again.")
    else:
        months, mval = [], {}
        for val, txt in opts:
            parts = txt.replace(",", "").split()
            if len(parts) >= 3:
                ml = f"{parts[0][:3].title()} {parts[-1]}"
                if ml not in mval:
                    months.append(ml)
                    mval[ml] = val

        st.caption("Net Investment \u00b7 INR Cr \u00b7 live from NSDL")

        bar_month = st.selectbox("Sector-wise flows \u2014 date", months, index=0, key="fpi_bar_month")
        with st.spinner("Loading\u2026"):
            bar_data = fpi_fetch((mval[bar_month],))
        cand = [p for p in bar_data if fpi_month(p) == bar_month]
        bar_period = sorted(cand, key=fpi_key)[-1] if cand else sorted(bar_data, key=fpi_key)[-1]
        sectors_all = [x for x in bar_data[bar_period] if x and x.lower() != "grand total"]

    

        vals = bar_data[bar_period]
        b = pd.DataFrame({"Sector": sectors_all,
                          "Value": [vals.get(s) for s in sectors_all]}).dropna()
        b["Sign"] = b["Value"].apply(lambda v: "Positive" if v >= 0 else "Negative")
        order = list(b.sort_values("Value", ascending=False)["Sector"])
        st.subheader(f"Sector-wise flows \u2014 {fpi_month(bar_period)}")
        yenc = alt.Y("Sector:N", sort=order, title=None)
        xenc = alt.X("Value:Q", title="Net Investment (INR Cr)",
                     scale=alt.Scale(reverse=True, nice=True))
        bars = alt.Chart(b).mark_bar().encode(
            x=xenc, y=yenc,
            color=alt.Color("Sign:N", scale=alt.Scale(domain=["Positive", "Negative"],
                                                      range=[NAVY, GREEN]), legend=None),
            tooltip=["Sector", "Value"])
        pos = alt.Chart(b[b["Value"] >= 0]).mark_text(
            align="right", baseline="middle", dx=-4, color="white", fontSize=10).encode(
            x=xenc, y=yenc, text=alt.Text("Value:Q", format=",.0f"))
        neg = alt.Chart(b[b["Value"] < 0]).mark_text(
            align="left", baseline="middle", dx=4, color="white", fontSize=10).encode(
            x=xenc, y=yenc, text=alt.Text("Value:Q", format=",.0f"))
        st.altair_chart((bars + pos + neg).properties(height=620).configure_axisY(labelLimit=400),
                        use_container_width=True)
        
        st.subheader("Sector trend")
        default_sec = "Metals & Mining" if "Metals & Mining" in sectors_all else sectors_all[0]
        sector = st.selectbox("Sector (line)", sectors_all,
                              index=sectors_all.index(default_sec), key="fpi_sector")
        pfrom = st.selectbox("From", months, index=min(9, len(months) - 1), key="fpi_from")
        pto = st.selectbox("To", months, index=0, key="fpi_to")

        i_from, i_to = months.index(pfrom), months.index(pto)
        lo, hi = min(i_from, i_to), max(i_from, i_to)
        need_vals = tuple(mval[m] for m in months[lo:hi + 1])

        with st.spinner("Loading range from NSDL\u2026"):
            rng = fpi_fetch(need_vals)

        keep = sorted((p for p in rng), key=fpi_key)
        ln = pd.DataFrame({"Period": [fpi_month(p) for p in keep],
                           "Value": [rng[p].get(sector) for p in keep]}).dropna()
        st.altair_chart(
            alt.Chart(ln).mark_line(point=alt.OverlayMarkDef(color=NAVY, size=55),
                                    strokeWidth=2, color=NAVY, interpolate="basis").encode(
                x=alt.X("Period:O", sort=list(ln["Period"]), title="Fortnight",
                        axis=alt.Axis(labelAngle=-40)),
                y=alt.Y("Value:Q", title="Net Investment (INR Cr)"),
                tooltip=["Period", "Value"],
            ).properties(height=340), use_container_width=True)

        st.download_button("CSV", ln.to_csv(index=False).encode(),
                           f"fpi_{sector[:12]}_{dt.date.today():%d%m%Y}.csv",
                           "text/csv", key="dl_fpi")
